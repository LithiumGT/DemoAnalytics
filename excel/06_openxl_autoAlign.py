from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime

wb = Workbook()
ws = wb.active

# Набор разнородных данных
ws.append([101, "Смартфон Core X", 89990.50, datetime(2026, 5, 26)])
ws.append([102, "Чехол силиконовый", 1200.00, datetime(2026, 5, 27)])

# Настройки выравнивания
align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")

# Алгоритм автоматического выравнивания
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4):
    for cell in row:
        val = cell.value

        if isinstance(val, (int, float)) and cell.column != 1:
            # Числа (исключая ID в первой колонке) — по правому краю
            cell.alignment = align_right
        elif isinstance(val, (datetime, int)):
            # Даты и ID — по центру
            cell.alignment = align_center
        else:
            # Строки и всё остальное — по левому краю
            cell.alignment = align_left

wb.save("aligned_report.xlsx")