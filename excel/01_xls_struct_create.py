from openpyxl import Workbook

# 1. Создаем объект Книги (в памяти)
wb = Workbook()

# 2. Получаем доступ к Листу
# При создании книги всегда создается один активный лист
ws = wb.active
ws.title = "Аналитика_2026"

# Создаем еще один лист в этой же книге
ws_archive = wb.create_sheet(title="Архив_Данных")

# 3. Работаем с конкретной Ячейкой
cell = ws["A1"]
cell.value = "Итоговая выручка"

# Проверяем связь объектов
print(type(wb))        # <class 'openpyxl.workbook.workbook.Workbook'>
print(type(ws))        # <class 'openpyxl.worksheet.worksheet.Worksheet'>
print(type(cell))      # <class 'openpyxl.cell.cell.Cell'>

wb.save("report.xlsx")
# Запись в конкретную точку
ws["B2"] = "ООО 'Вектор'"

# Получение среза ячеек (вернет кортеж кортежей ячеек)
cell_range = ws["A4:C10"]

ws.cell(row=5, column=2, value="Динамическое значение")

from openpyxl.utils import get_column_letter, column_index_from_string

# Получить букву колонки по ее номеру
print(get_column_letter(2))  # Выведет: 'B'
print(get_column_letter(28)) # Выведет: 'AB'

# Получить номер колонки по ее букве
print(column_index_from_string("C"))  # Выведет: 3