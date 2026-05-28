from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

wb = Workbook()
ws = wb.active
ws.title = "Фиксация_Областей"

# 1. Генерируем большую матрицу данных (10 колонок x 50 строк)
# Добавляем шапку
ws.append(["ID", "Сотрудник"] + [f"Месяц_{i}" for i in range(1, 9)])

# Добавляем строки с данными
for r in range(1, 50):
    ws.append([f"EMP{r:03}", f"Иван Иванов {r}"] + [r * 1000 for _ in range(8)])

# 2. Немного стилизуем шапку и первую колонку для визуального акцента
header_fill = PatternFill(fill_type="solid", start_color="1F497D")
header_font = Font(color="FFFFFF", bold=True)
id_font = Font(bold=True)

# Красим шапку (строка 1)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font

# Делаем жирным ID (колонка A)
for row in ws.iter_rows(min_row=2, max_row=50, min_col=1, max_col=1):
    for cell in row:
        cell.font = id_font

# 3. НАСТРОЙКА ЗАКРЕПЛЕНИЯ ОБЛАСТЕЙ (Freeze Panes)
# Указываем ячейку B2.
# Всё, что ВЫШЕ строки 2 (строка 1) и ЛЕВЕЕ колонки B (колонка A), заморозится.
ws.freeze_panes = "B2"

wb.save("freeze_panes_report.xlsx")