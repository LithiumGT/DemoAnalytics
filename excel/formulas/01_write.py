from openpyxl import Workbook

# Создаем книгу и лист
wb = Workbook()
ws = wb.active

# Записываем тестовые данные
ws["A3"] = 100
ws["A4"] = 200
ws["A5"] = 300

# Записываем формулу
ws["A6"] = "=SUM(A3:A5)"

# Сохраняем файл.
wb.save("test_formulas.xlsx")