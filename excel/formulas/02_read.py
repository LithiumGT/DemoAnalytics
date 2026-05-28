import pandas as pd
from openpyxl import load_workbook

print("--- ПОПЫТКА 1: openpyxl без флага data_only ---")
wb_raw = load_workbook("test_formulas.xlsx", data_only=False)
ws_raw = wb_raw.active
print(f"Значение в A6: {ws_raw['A6'].value}")
# Результат: строка "=SUM(A3:A5)"
# openpyxl просто честно прочитал то, что было записано: текст формулы.


print("\n--- ПОПЫТКА 2: openpyxl с флагом data_only=True ---")
wb_data = load_workbook("test_formulas.xlsx", data_only=True)
ws_data = wb_data.active
print(f"Значение в A6: {ws_data['A6'].value}")
# Результат: None

print("\n--- ПОПЫТКА 3: pandas (read_excel) ---")
# pandas под капотом использует тот же openpyxl в режиме чтения значений
# Попытка 3 (Исправленная)
df = pd.read_excel("test_formulas.xlsx", header=None)

# Берём самую последнюю строку (индекс -1), первая колонка (0)
print(f"Значение в последней строке (ячейка A6): {df.iloc[-1, 0]}")