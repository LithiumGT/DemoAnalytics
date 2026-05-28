from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = Workbook()
ws = wb.active

# Подготовка данных
ws.append(["ID", "Товар", "Цена"])
ws.append([1, "Ноутбук", 75000])

# Стили для шапки
h_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
h_fill = PatternFill(fill_type="solid", start_color="4F81BD")
h_align = Alignment(horizontal="center")
thin_border = Border(bottom=Side(style="medium", color="000000"))

# Стилизуем всю первую строку (шапку) динамически
for col_idx in range(1, 4):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = h_font
    cell.fill = h_fill
    cell.alignment = h_align
    cell.border = thin_border

wb.save("styled_report.xlsx")