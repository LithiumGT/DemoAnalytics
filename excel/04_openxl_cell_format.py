from openpyxl import Workbook
from datetime import datetime

wb = Workbook()
ws = wb.active

# 1. Денежный формат с разделителями тысяч и знаком рубля
ws["A1"] = 1250500.5
ws["A1"].number_format = '#,##0.00 ₽'  # Отобразится: 1 250 500,50 ₽

# 2. Процентный формат
ws["B1"] = 0.154
ws["B1"].number_format = '0.0%'        # Отобразится: 15.4%

# 3. Формат даты
ws["C1"] = datetime(2026, 5, 26)
ws["C1"].number_format = 'YYYY-MM-DD'  # Отобразится: 2026-05-26

wb.save("formatted_report.xlsx")